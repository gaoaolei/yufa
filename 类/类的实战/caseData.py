import openpyxl
from .caseParam import CaseParam

# 黄灿的方法
class CaseData(object):
    def __init__(self, wbname):
        self.wbname = wbname
        self.data = []
        self.getCaseData()

    def getCaseData(self):
        wb = openpyxl.load_workbook(self.wbname)
        ws = wb[wb.sheetnames[0]]
        max_row = ws.max_row + 1
        max_col = ws.max_column + 1
        for row in range(2, max_row):
            caseParam = CaseParam()
            for col in range(1, max_col):
                param_name = ws.cell(1, col).value
                if param_name in caseParam.all_param:
                    value = ws.cell(row, col).value
                    if value is not None:
                        value = str(value).strip()
                    setattr(caseParam, param_name, value)
            self.data.append(caseParam)


# 我的优化  思路：直接继承list，省去data属性，更方便
class CaseData1(list):
    def __init__(self, wbname):
        super(CaseData1, self).__init__()  # list的初始化可传空
        self.wbname = wbname
        # self.data = []
        self.getCaseData()

    def getCaseData(self):
        wb = openpyxl.load_workbook(self.wbname)
        ws = wb[wb.sheetnames[0]]
        max_row = ws.max_row + 1
        max_col = ws.max_column + 1
        for row in range(2, max_row):
            caseParam = CaseParam()
            for col in range(1, max_col):
                param_name = ws.cell(1, col).value
                if param_name in caseParam.all_param:
                    value = ws.cell(row, col).value
                    if value is not None:
                        value = str(value).strip()
                    setattr(caseParam, param_name, value)
            self.append(caseParam)

if __name__ == "__main__":
    a = CaseData('a.xlsx')
    a1 = CaseData1('a.xlsx')
    print(a, type(a))
    print(a.data, type(a.data))
    print(a1, (type(a1)))
    print(isinstance(a1, list))
