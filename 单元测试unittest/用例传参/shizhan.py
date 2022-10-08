
import unittest
import HTMLTestRunner
import openpyxl

users = [{'user': 'python', 'password': '123456'}]


def register(username, password1, password2):
    # 注册功能
    for user in users:  # 遍历出所有账号，判断账号是否存在
        if username == user['user']:
            # 账号存在
            return {"code": 0, "msg": "该账户已存在"}
    else:
        if password1 != password2:
            # 两次密码不一致
            return {"code": 0, "msg": "两次密码不一致"}
        else:
            # 账号不存在 密码重复，判断账号密码长度是否在 6-18位之间
            if 6 <= len(username) >= 6 and 6 <= len(password1) <= 18:
                # 注册账号
                users.append({'user': username, 'password': password2})
                return {"code": 1, "msg": "注册成功"}
            else:
                # 账号密码长度不对，注册失败
                return {"code": 0, "msg": "账号和密码必须在6-18位之间"}


class Case:
    # 这个类用来存储用例
    def __init__(self, attrs):
        """
        初始化用例
        :param attrs: zip类型-->[(key,value),(key1,value1)....]
        """
        for item in attrs:
            setattr(self, item[0], item[1])


class ReadExcel(object):
    """
    读取excel数据
    """

    def __init__(self, file_name, sheet_name):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        :param sheet_name: 表单名字  -->  str
        """
        self.file_name = file_name
        # 打开工作簿
        self.wb = openpyxl.load_workbook(file_name)
        # 选择表单
        self.sheet = self.wb[sheet_name]

    def __del__(self):
        # 特殊的方法，在对象销毁的之后执行
        # 关闭文件
        self.wb.close()

    def read_data_obj(self, list1):
        # list1 参数为一个列表，传入的是指定读取数据的列,比如[1,2,3]
        # 每一行[1,3,5]列的数据，读取出来就作为一条测试用例，放在对象的属性中
        # 所有的用例对象放在列表中并且进行返回
        max_r = self.sheet.max_row
        # 定义一个空列表，用来存放所有用例
        cases = []
        titles = []  # 定义一个空列表，用来存放表头
        # 遍历出所有的行
        for row in range(1, max_r + 1):
            # 判断是否是第一行
            if row != 1:
                case_data = []  # 定义一个空列表，用来存放该行的数据
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                # 将该条数据和表头进行打包组合，
                case = list(zip(titles, case_data))
                # 将一条用例存入一个对象中（每一列对应对象的一个属性）
                # case_obj=Case()
                # for i in case:
                #     setattr(case_obj,i[0],i[1]) # 这一部分已经写到类定义的地方去了
                # 不用上面注释的，直接像下面这样传入就可以了
                case_obj = Case(case)
                cases.append(case_obj)
            else:
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        return cases  # 所有的用例对象放在列表中并且进行返回


class RegisterTestCase(unittest.TestCase):
    def __init__(self, expected, data, methodName):
        self.expected = expected
        self.data = data
        super(RegisterTestCase, self).__init__(methodName)

    def test_register(self):
        print("--------------")
        print(users)
        res = register(*self.data)
        print(users)
        try:
            self.assertEqual(self.expected, res)
        except AssertionError as e:
            print('该用例测试未通过')
            raise e
        else:
            print('该用例测试通过')

    def test_001(self):
        self.assertTrue(0)

# 创建测试集合
suite = unittest.TestSuite()  # 报错就是因为这个没有加括号
# 读取excel中的数据
r = ReadExcel('cases.xlsx', 'Sheet1')
cases = r.read_data_obj([2, 3])
# 把每条用例遍历出来
for case in cases:
    suite.addTest(RegisterTestCase(eval(case.expected), eval(case.data), "test_register"))
    suite.addTest(RegisterTestCase(eval(case.expected), eval(case.data), "test_001"))
# 执行测试集合
with open('old_reports.html', 'wb') as f:
    runner = HTMLTestRunner.HTMLTestRunner(
        stream=f,
        verbosity=2,
        title='old_report',
        description='旧的测试报告')
    runner.run(suite)
