import openpyxl
import os
__author__ = "Administrator"

from airtest.core.api import *
from poco.drivers.android.uiautomation import AndroidUiautomationPoco

poco = AndroidUiautomationPoco(use_airtest_input=True, screenshot_each_action=False)
import unittest
from airtest.cli.parser import cli_setup
from airtest.report.report import simple_report
import re

auto_setup(__file__)


class HandleExcel:

    def __init__(self, filename, sheetname):
        """
        初始化
        :param filename:文件名
        :param sheetname:表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        # 创建工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 创建表单对象
        sh = wb[self.sheetname]
        # 按行读取内容
        data = list(sh.rows)

        # 创建空列表存放用例内容
        cases_data = []
        # # 获取表头数据，并保存在列表中
        # title = []
        # for item in data[0]:
        #     title.append(item.value)
        #
        # # 遍历读取用例内容
        for item in data:
            case_data = []
            for i in item:
                if i.value is None:
                    case_data.append('')
                else:
                    case_data.append(str(i.value))
        #     # 将表头和每行用例内容进行打包
        #     case = dict(zip(title, case_data))
        #     # 将打包后的用例内容储存在用例列表中
            cases_data.append(case_data)
        return cases_data

    def write_excel(self, row, column, value):
        # 创建工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 创建表单对象
        sh = wb[self.sheetname]
        # 写入内容
        sh.cell(row=row, column=column, value=value)
        # 保存文件
        wb.save(self.filename)

if __name__ == '__main__':
    excel = HandleExcel( '人人驾统一全量司机11月.xlsx', 'Sheet')
    cases = excel.read_excel()
    # print(cases)
    for i in enumerate(cases):
        # print(i)
        # if i[1][3] != '人人驾段麒麟7队':
        #     continue
        # print(type(i[0]))
        if i[0] == 0:
            continue
        if i[1][6]:
            continue
        name = i[1][2]
        mobile = i[1][5]
        print(name)
        print(mobile)
        # ---------------------
        poco('com.tencent.mm:id/plus_icon').click()
        poco(text='添加朋友').click()
        poco(text='账号/手机号').click()
        sleep(0.5)
        text(mobile, enter=True)
        poco(textMatches='^搜索.*').click()
        if poco(text='该用户不存在'):
            swipe((1, 1000), (300, 1000))
            sleep(1)
            swipe((1, 1000), (300, 1000))
            excel.write_excel(i[0] + 1, 7, '该用户不存在')
        elif not poco(text='添加到通讯录').exists():
            swipe((1, 1000), (300, 1000))
            sleep(0.5)
            swipe((1, 1000), (300, 1000))
            sleep(0.5)
            swipe((1, 1000), (300, 1000))
        else:
            poco(text='添加到通讯录').click()
            # sleep(0.2)
            # poco('com.tencent.mm:id/j0z').click()
            # poco('com.tencent.mm:id/j0z').set_text("")
            # text(name)
            # # 点击标签
            # poco('com.tencent.mm:id/bo0').click()
            # poco(text='选择或搜索标签').click()
            # poco(text='选择或搜索标签').set_text("")
            # text('7队')
            sleep(1)
            poco('com.tencent.mm:id/m9y').set_text('解封账号，平移车证')
            sleep(0.3)
            # poco(text='保存').click()
            poco(text='发送').click()
            sleep(2)
            swipe((1, 1000), (300, 1000))
            sleep(0.5)
            swipe((1, 1000), (300, 1000))
            sleep(1)
            swipe((1, 1000), (300, 1000))
            excel.write_excel(i[0]+1,7,'已添加待通过')
