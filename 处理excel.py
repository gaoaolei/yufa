import openpyxl
import os

class HandleExcel:

    def __init__(self, filename, sheetname):
        """
        初始化
        :param filename:文件名
        :param sheetname:表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        # 创建工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 创建表单对象
        sh = wb[self.sheetname]
        # 按行读取内容
        data = list(sh.rows)

        # 创建空列表存放用例内容
        cases_data = []
        # # 获取表头数据，并保存在列表中
        # title = []
        # for item in data[0]:
        #     title.append(item.value)
        #
        # # 遍历读取用例内容
        for item in data:
            case_data = []
            for i in item:
                if i.value is None:
                    case_data.append('')
                else:
                    case_data.append(str(i.value))
        #     # 将表头和每行用例内容进行打包
        #     case = dict(zip(title, case_data))
        #     # 将打包后的用例内容储存在用例列表中
            cases_data.append(case_data)
        return cases_data

    def write_excel(self, row, column, value):
        # 创建工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 创建表单对象
        sh = wb[self.sheetname]
        # 写入内容
        sh.cell(row=row, column=column, value=value)
        # 保存文件
        wb.save(self.filename)

if __name__ == '__main__':
    excel = HandleExcel( 'a.xlsx', 'ad')
    cases = excel.read_excel()
    # print(cases)
    for i in enumerate(cases):
        # print(i[0])
        # print(i[1])
        print(f'ct.campaign.configInfo[{i[0]}].configName = {i[1][1]}')
        print(f'ct.campaign.configInfo[{i[0]}].configPrizeId = {i[1][2]}')
        print(f'ct.campaign.configInfo[{i[0]}].configLogoUrl = {i[1][9]}')
        print(f'ct.campaign.configInfo[{i[0]}].configLinkUrl = {i[1][10]}')